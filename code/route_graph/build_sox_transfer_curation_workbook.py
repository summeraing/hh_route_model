from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path.cwd() / "TOP_JOURNAL_REBUILD_ANALYSES_v1" / "iMETA_ROUTE_B_REBUILD_20260630"
CASE = ROOT / "15_iMETA_ROUTE_B_DUAL_CASE_TRANSFER_20260701" / "01_HGT_SOX_CASE"
OUT = CASE / "SOX_HGT_transfer_curation_workbook_v1.xlsx"


def read_csv(rel: str) -> pd.DataFrame:
    return pd.read_csv(CASE / rel)


def autosize_and_style(path: Path) -> None:
    wb = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        for col_idx, col_cells in enumerate(ws.columns, start=1):
            max_len = 10
            for cell in col_cells[:200]:
                if cell.value is None:
                    continue
                max_len = max(max_len, min(len(str(cell.value)), 55))
            ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 2
    wb.save(path)


def main() -> None:
    instructions = pd.DataFrame(
        [
            {
                "step": 1,
                "action": "Read schema and coding manual before filling evidence_units.",
                "notes": "Rows must be traceable public evidence units, not free-form impressions.",
            },
            {
                "step": 2,
                "action": "Fill source_registry first.",
                "notes": "Each source_id used in evidence_units must appear in source_registry.",
            },
            {
                "step": 3,
                "action": "Fill evidence_units with donor_class, functional_class and dependency_group.",
                "notes": "Use uncertain/support instead of forcing ambiguous rows into strict route labels.",
            },
            {
                "step": 4,
                "action": "Run scripts/run_route_graph_case.py and inspect preflight_report.",
                "notes": "Do not use the SOX case in the manuscript until the preflight gates pass or the failure is explicitly framed as a boundary case.",
            },
        ]
    )

    schema = read_csv("schema/route_graph_input_schema.csv")
    source_registry = read_csv("source_curation/sox_candidate_source_registry_v1.csv")
    backlog = read_csv("source_curation/sox_extraction_backlog_v1.csv")
    seed = read_csv("curated_seed/sox_evidence_units_seed_v1.csv")
    evidence_template = read_csv("input_templates/sox_evidence_units_template.csv")
    blank_rows = pd.concat([evidence_template.iloc[0:0], pd.DataFrame([{} for _ in range(150)])], ignore_index=True)

    with pd.ExcelWriter(OUT, engine="openpyxl") as writer:
        instructions.to_excel(writer, sheet_name="Instructions", index=False)
        schema.to_excel(writer, sheet_name="Input_schema", index=False)
        source_registry.to_excel(writer, sheet_name="Candidate_sources", index=False)
        backlog.to_excel(writer, sheet_name="Extraction_backlog", index=False)
        seed.to_excel(writer, sheet_name="Seed_units_NOT_FINAL", index=False)
        blank_rows.to_excel(writer, sheet_name="Evidence_units_FILL_HERE", index=False)

    autosize_and_style(OUT)
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    main()
